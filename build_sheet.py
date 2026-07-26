# Genera el sheet de pipeline de cold outreach para Kurogrid.
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Junto al script. Antes apuntaba a un directorio temporal de la sesión en que
# se escribió esto: el archivo se generaba fuera del repo y el de acá quedaba
# viejo sin que nada avisara (y ese temporal ya no existe).
OUT = Path(__file__).resolve().parent / "kurogrid-outreach-pipeline.xlsx"

INK = "1A1A1A"
ACCENT = "9EF01A"  # lime kurogrid
HEADER_FILL = PatternFill("solid", fgColor=INK)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ESTADOS = [
    "Por contactar",
    "Contactado",
    "Follow-up enviado",
    "Respondió",
    "En conversación",
    "Propuesta enviada",
    "Cerrado ganado",
    "Cerrado perdido",
    "No responde",
]
ESTADO_COLORS = {
    "Por contactar": "F2F2F2",
    "Contactado": "DDEBF7",
    "Follow-up enviado": "BDD7EE",
    "Respondió": "FFF2CC",
    "En conversación": "FFE699",
    "Propuesta enviada": "FFD966",
    "Cerrado ganado": "C6EFCE",
    "Cerrado perdido": "FFC7CE",
    "No responde": "D9D9D9",
}
CANALES = ["Email", "LinkedIn", "WhatsApp", "Instagram", "Referido", "Llamada"]
PLANES = [
    "Presencia S/199",
    "Empresa S/449",
    "Empresa+ S/649",
    "Sistemas S/999+",
    "Proyecto puntual",
    "Sin definir",
]

wb = Workbook()

# ─── Hoja 1: Pipeline ───────────────────────────────────────
ws = wb.active
ws.title = "Pipeline"

cols = [
    ("Empresa", 24),
    ("Contacto", 18),
    ("Cargo", 16),
    ("Rubro", 16),
    ("Canal", 13),
    ("WhatsApp", 15),
    ("Email", 26),
    ("Web actual", 26),
    ("Gancho / nota personalizada", 38),
    ("1er contacto", 13),
    ("# Follow-ups", 12),
    ("Próximo seguimiento", 18),
    ("Estado", 19),
    ("Plan de interés", 17),
    ("Resultado / motivo", 30),
]
for i, (name, width) in enumerate(cols, start=1):
    c = ws.cell(row=1, column=i, value=name)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

N = 300  # filas con validación/formato listas para usar

dv_estado = DataValidation(type="list", formula1='"' + ",".join(ESTADOS) + '"', allow_blank=True)
dv_canal = DataValidation(type="list", formula1='"' + ",".join(CANALES) + '"', allow_blank=True)
dv_plan = DataValidation(type="list", formula1='"' + ",".join(PLANES) + '"', allow_blank=True)
ws.add_data_validation(dv_estado)
ws.add_data_validation(dv_canal)
ws.add_data_validation(dv_plan)
dv_canal.add(f"E2:E{N}")
dv_estado.add(f"M2:M{N}")
dv_plan.add(f"N2:N{N}")

# Colores por estado (fila completa suave en la celda de estado)
for estado, color in ESTADO_COLORS.items():
    ws.conditional_formatting.add(
        f"M2:M{N}",
        FormulaRule(formula=[f'$M2="{estado}"'], fill=PatternFill("solid", fgColor=color)),
    )
# Resaltar seguimientos vencidos (fecha próxima <= hoy y no cerrado)
ws.conditional_formatting.add(
    f"L2:L{N}",
    FormulaRule(
        formula=['AND($L2<>"",$L2<=TODAY(),$M2<>"Cerrado ganado",$M2<>"Cerrado perdido")'],
        fill=PatternFill("solid", fgColor="FFC7CE"),
        font=Font(bold=True, color="9C0006"),
    ),
)

# Fila de ejemplo
ejemplo = [
    "Clínica San Martín (ejemplo)", "Dra. Rojas", "Administradora", "Salud",
    "Email", "51 999 888 777", "admin@clinicasm.pe", "clinicasm.pe (lenta, sin citas online)",
    "Su web no carga en móvil; ofrecer Empresa+ con módulo de citas",
    "2026-07-20", 1, "2026-07-23", "Contactado", "Empresa+ S/649", "",
]
for i, v in enumerate(ejemplo, start=1):
    ws.cell(row=2, column=i, value=v)

for row in ws.iter_rows(min_row=2, max_row=N, max_col=len(cols)):
    for c in row:
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
for r in (f"J2:J{N}", f"L2:L{N}"):
    pass  # formato de fecha lo maneja Sheets al importar

# ─── Hoja 2: Links UTM ──────────────────────────────────────
ws2 = wb.create_sheet("Links UTM")
ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 78
ws2.column_dimensions["D"].width = 46

ws2["A1"] = "Canal"
ws2["B1"] = "Página destino"
ws2["C1"] = "Link para pegar en el mensaje"
ws2["D1"] = "Cuándo usarlo"
for cell in ("A1", "B1", "C1", "D1"):
    ws2[cell].fill = HEADER_FILL
    ws2[cell].font = HEADER_FONT

BASE = "https://kurogrid.com"
# /promo es la landing de una sola oferta: no tiene menú ni rutas para irse a
# otro lado, así que es la mejor caída para tráfico frío, que llega sin
# contexto y se pierde fácil. /web es más amplia y es la que Google indexa:
# se reserva para quien llega por su cuenta o ya te conoce.
links = [
    ("Email", "/promo", f"{BASE}/promo?utm_source=outreach&utm_medium=email&utm_campaign=cold-v1", "Email frío a empresas que necesitan web. Caída por defecto del cold approach."),
    ("Email", "/software", f"{BASE}/software?utm_source=outreach&utm_medium=email&utm_campaign=cold-v1", "Email frío a empresas con procesos manuales (software a medida)"),
    ("LinkedIn", "/promo", f"{BASE}/promo?utm_source=linkedin&utm_medium=dm&utm_campaign=cold-v1", "DM de LinkedIn después de conectar"),
    ("LinkedIn", "/software", f"{BASE}/software?utm_source=linkedin&utm_medium=dm&utm_campaign=cold-v1", "DM de LinkedIn, perfil técnico/operaciones"),
    ("Instagram", "/promo", f"{BASE}/promo?utm_source=instagram&utm_medium=dm&utm_campaign=cold-v1", "DM a negocios activos en IG sin web decente"),
    ("WhatsApp", "/web", f"{BASE}/web?utm_source=whatsapp&utm_medium=dm&utm_campaign=referidos", "Solo referidos o quien ya respondió por otro canal"),
    ("Referido", "/web", f"{BASE}/web?utm_source=referido&utm_medium=word-of-mouth&utm_campaign=organico", "Cuando un cliente o contacto te recomienda"),
    ("WhatsApp", "/promo", f"{BASE}/promo?utm_source=whatsapp&utm_medium=dm&utm_campaign=cold-v1", "WhatsApp frío, cuando conseguiste el número por otra vía"),
]
for i, row in enumerate(links, start=2):
    for j, v in enumerate(row, start=1):
        c = ws2.cell(row=i, column=j, value=v)
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)

notes = [
    "",
    "Cómo funciona: el visitante llega con el link y los UTMs quedan guardados en su sesión.",
    "Desde ahí viajan a dos lugares distintos del portal:",
    "  1. Si clickea cualquier botón de WhatsApp -> site_demand_clicks. Ahí ves qué canal genera clics.",
    "  2. Si llena el formulario -> la oportunidad en Oportunidades queda con su canal y su tanda.",
    "",
    "Incluso si abandona el formulario a medio camino: mientras haya llegado a poner nombre, empresa",
    "y WhatsApp, la oportunidad se guarda igual, etiquetada '— incompleto' y diciendo en qué paso se fue.",
    "Esos son los que más conviene llamar rápido: dejaron el teléfono pero no terminaron de pedirte nada.",
    "",
    "A dónde mandar: /promo para tráfico frío (una sola oferta, sin menú para irse a otro lado).",
    "/web para referidos y para quien ya te conoce; es además la que Google indexa.",
    "",
    "utm_campaign NO se cambia por mes. Se cambia cuando cambias algo que quieres medir aparte:",
    "reescribiste el pitch, cambiaste el tipo de empresa al que apuntas, o probaste un enfoque nuevo.",
    "Ese día pasas a cold-v2. Si mandas el mismo mensaje tres meses, cold-v1 se queda tres meses.",
    "",
    "Con pocos leads, partir por mes es peor: cubetas de 2 o 3 no distinguen una mejora de una",
    "casualidad. Conviene acumular. Y aunque nunca lo cambies, utm_source y utm_medium siguen",
    "diciéndote si LinkedIn rinde más que el correo; lo único que pierdes es el antes/después.",
    "",
    "No inventes más valores de utm_source de los necesarios: un valor por canal, siempre igual.",
]
start = len(links) + 3
for k, t in enumerate(notes):
    c = ws2.cell(row=start + k, column=1, value=t)
    c.font = Font(italic=True, color="666666", size=9)

# ─── Hoja 3: Guía rápida ────────────────────────────────────
ws3 = wb.create_sheet("Guía")
ws3.column_dimensions["A"].width = 24
ws3.column_dimensions["B"].width = 90

guia = [
    ("FLUJO DE ESTADOS", ""),
    ("Por contactar", "Prospecto identificado, aún sin mensaje enviado."),
    ("Contactado", "Primer mensaje enviado. Anota la fecha en '1er contacto' y pon 'Próximo seguimiento' a +3 días."),
    ("Follow-up enviado", "Mandaste un recordatorio. Suma 1 en '# Follow-ups'. Máximo 2-3 follow-ups y luego 'No responde'."),
    ("Respondió", "Contestó cualquier cosa. Objetivo cumplido del mensaje frío: ahora conversa, no vendas todavía."),
    ("En conversación", "Hay interés real, están intercambiando mensajes o agendaron llamada."),
    ("Propuesta enviada", "Le mandaste plan y precio. Seguimiento a +2 días si no contesta."),
    ("Cerrado ganado", "Aceptó. Pasa al onboarding del portal."),
    ("Cerrado perdido", "Dijo que no. Anota el motivo en 'Resultado' — eso te dice qué ajustar del pitch."),
    ("No responde", "2-3 follow-ups sin respuesta. No insistas más; puedes reintentar en 2-3 meses."),
    ("", ""),
    ("CADENCIA SUGERIDA", ""),
    ("Día 0", "Primer mensaje: corto, personalizado (usa la columna 'Gancho'), una sola pregunta al final."),
    ("Día 3", "Follow-up 1: una línea, aporta algo nuevo (ej. algo que viste de su negocio), no repitas el pitch."),
    ("Día 7-8", "Follow-up 2: último intento, cierre amable ('si no es el momento, sin problema')."),
    ("", ""),
    ("REGLAS DE ORO", ""),
    ("Personalización", "El 'Gancho' es obligatorio antes de contactar: algo específico de SU negocio (web lenta, sin web, sin citas online, reseñas buenas pero cero presencia)."),
    ("Volumen", "Mejor 10 mensajes personalizados al día que 50 genéricos. WhatsApp frío con cuidado: puede quemar el número."),
    ("El link", "Usa siempre los links de la hoja 'Links UTM' según el canal, nunca el dominio pelado."),
    ("Revisión semanal", "Filtra por 'Próximo seguimiento' vencido (se pinta rojo) cada lunes y ejecuta."),
]
for i, (a, b) in enumerate(guia, start=1):
    ca = ws3.cell(row=i, column=1, value=a)
    cb = ws3.cell(row=i, column=2, value=b)
    cb.alignment = Alignment(vertical="center", wrap_text=True)
    if b == "" and a:
        ca.font = Font(bold=True, size=11)
        ca.fill = PatternFill("solid", fgColor="F2F2F2")
    else:
        ca.font = Font(bold=True, size=10)

wb.save(OUT)
print("saved", OUT)
